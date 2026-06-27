import io
import re
import zipfile
import unicodedata
import xml.etree.ElementTree as ET

import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

APP_TITLE = "BRP: Excel imprimible por docente"
TEMPLATE_PATH = "plantilla.xlsx"
DATA_ROW = 160
PRINT_AREA = "A162:C193"
MAX_REGISTROS = 300


def normalize(text):
    return str(text).strip().lower() if text is not None else ""


def strip_accents(text):
    text = "" if text is None else str(text).strip()
    return "".join(
        c for c in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(c)
    )


def sort_key(text):
    return strip_accents(text).lower()


def clean_sheet_name(text):
    text = "" if text is None else str(text).strip()
    text = re.sub(r'[:\\/?*\[\]]', " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def safe_sheet_title(name, existing):
    name = clean_sheet_name(name) or "Hoja"
    base = name[:31]
    title = base
    i = 1

    while title in existing:
        suffix = f"_{i}"
        title = base[:31 - len(suffix)] + suffix
        i += 1

    return title[:31]


def col_letters_to_number(cell_ref):
    letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
    number = 0
    for ch in letters:
        number = number * 26 + (ord(ch) - ord("A") + 1)
    return number


def read_shared_strings(zf):
    path = "xl/sharedStrings.xml"
    if path not in zf.namelist():
        return []

    root = ET.fromstring(zf.read(path))
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    strings = []
    for si in root.findall("a:si", ns):
        texts = []
        for t in si.findall(".//a:t", ns):
            texts.append(t.text or "")
        strings.append("".join(texts))

    return strings


def get_first_sheet_path(zf):
    workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    ns_main = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    ns_rel = {
        "rel": "http://schemas.openxmlformats.org/package/2006/relationships"
    }

    first_sheet = workbook_xml.find("a:sheets/a:sheet", ns_main)
    rel_id = first_sheet.attrib[
        "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    ]

    for rel in rels_xml.findall("rel:Relationship", ns_rel):
        if rel.attrib.get("Id") == rel_id:
            target = rel.attrib.get("Target")
            if target.startswith("/"):
                return target.lstrip("/")
            return "xl/" + target.lstrip("/")

    raise ValueError("No se pudo encontrar la hoja principal del archivo.")


def read_xlsx_values(base_bytes):
    """
    Lee valores desde el archivo .xlsx sin cargar estilos.
    Esto evita errores de openpyxl con formatos internos defectuosos.
    """
    rows = []

    with zipfile.ZipFile(io.BytesIO(base_bytes)) as zf:
        shared_strings = read_shared_strings(zf)
        sheet_path = get_first_sheet_path(zf)

        root = ET.fromstring(zf.read(sheet_path))
        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        for row in root.findall(".//a:sheetData/a:row", ns):
            values = {}

            for cell in row.findall("a:c", ns):
                cell_ref = cell.attrib.get("r", "")
                col_num = col_letters_to_number(cell_ref)
                cell_type = cell.attrib.get("t")
                value_node = cell.find("a:v", ns)

                value = ""

                if value_node is not None:
                    raw_value = value_node.text or ""

                    if cell_type == "s":
                        idx = int(raw_value)
                        value = shared_strings[idx] if idx < len(shared_strings) else ""
                    else:
                        value = raw_value

                # inline strings
                if cell_type == "inlineStr":
                    text_node = cell.find(".//a:t", ns)
                    value = text_node.text if text_node is not None else ""

                values[col_num] = value

            if values:
                max_col = max(values.keys())
                rows.append([values.get(c, "") for c in range(1, max_col + 1)])

    return rows


def leer_registros(base_bytes):
    rows = read_xlsx_values(base_bytes)

    if not rows:
        raise ValueError("El archivo está vacío o no pudo leerse.")

    headers = [normalize(h) for h in rows[0]]
    col = {headers[i]: i for i in range(len(headers)) if headers[i]}

    if "rut (docente)" not in col:
        raise ValueError(
            "No se encontró la columna 'RUT (Docente)'. "
            "Suba el archivo BRP mensual correcto."
        )

    rut_idx = col.get("rut (docente)")
    nom_idx = col.get("nombres (docente)")
    ap1_idx = col.get("primer apellido (docente)")
    ap2_idx = col.get("segundo apellido (docente)")
    rbd_idx = col.get("rbd (establecimiento)")
    periodo_idx = col.get("período") or col.get("periodo")
    mes_idx = col.get("mes")
    anio_idx = col.get("año") or col.get("ano")

    registros = []
    rbd_detectado = ""
    periodo_detectado = ""

    for row in rows[1:]:
        rut = row[rut_idx] if rut_idx < len(row) else ""

        if rut in (None, ""):
            continue

        if not rbd_detectado and rbd_idx is not None and rbd_idx < len(row):
            rbd_detectado = row[rbd_idx] or ""

        if not periodo_detectado:
            if periodo_idx is not None and periodo_idx < len(row):
                periodo_detectado = row[periodo_idx] or ""
            elif mes_idx is not None and anio_idx is not None:
                mes = row[mes_idx] if mes_idx < len(row) else ""
                anio = row[anio_idx] if anio_idx < len(row) else ""
                periodo_detectado = f"{mes} {anio}".strip()

        ap1 = row[ap1_idx] if ap1_idx is not None and ap1_idx < len(row) else ""
        ap2 = row[ap2_idx] if ap2_idx is not None and ap2_idx < len(row) else ""
        nom = row[nom_idx] if nom_idx is not None and nom_idx < len(row) else ""

        registros.append(
            {
                "rut": rut,
                "ap1": ap1,
                "ap2": ap2,
                "nom": nom,
                "row_values": row,
            }
        )

    if not registros:
        raise ValueError("El archivo no contiene registros con RUT.")

    registros.sort(
        key=lambda x: (
            sort_key(x["ap1"]),
            sort_key(x["ap2"]),
            sort_key(x["nom"]),
            str(x["rut"]).strip(),
        )
    )

    return registros, rbd_detectado, periodo_detectado


def procesar(base_bytes):
    registros, _, _ = leer_registros(base_bytes)

    if len(registros) > MAX_REGISTROS:
        raise ValueError(
            f"Se detectaron {len(registros)} registros. "
            "Este archivo parece ser histórico y no mensual. "
            "Por favor suba el archivo BRP mensual correcto."
        )

    wb_out = openpyxl.load_workbook(TEMPLATE_PATH, data_only=False)
    tmpl = wb_out.active

    wb_out.calculation.fullCalcOnLoad = True
    wb_out.calculation.forceFullCalc = True
    wb_out.calculation.calcMode = "auto"

    progress = st.progress(0)

    for idx, reg in enumerate(registros, start=1):
        ws = wb_out.copy_worksheet(tmpl)

        ws.title = safe_sheet_title(
            f'{reg["ap1"]}_{reg["ap2"]}_{reg["rut"]}',
            set(wb_out.sheetnames),
        )

        for c, val in enumerate(reg["row_values"], start=1):
            ws.cell(DATA_ROW, c).value = val

        ws.print_area = PRINT_AREA
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.paperSize = 9

        ws.page_margins = PageMargins(
            left=0.25,
            right=0.25,
            top=0.5,
            bottom=0.5,
            header=0.3,
            footer=0.3,
        )

        for rr in range(1, 162):
            ws.row_dimensions[rr].hidden = True

        for cc in range(4, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(cc)].hidden = True

        ws.sheet_view.showGridLines = False
        ws.sheet_view.selection[0].activeCell = "A162"
        ws.sheet_view.selection[0].sqref = "A162"
        ws.sheet_view.topLeftCell = "A162"

        progress.progress(idx / len(registros))

    wb_out.remove(tmpl)

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)

    return output.getvalue(), len(registros)


st.set_page_config(page_title=APP_TITLE, page_icon="🧾")

st.title("🧾 " + APP_TITLE)

st.markdown(
    """
**Instrucciones:**

1. Descargue el archivo BRP mensual.
2. Súbalo aquí sin modificarlo.
3. Revise que el período y cantidad de docentes sean correctos.
4. Descargue el Excel final listo para imprimir.

✅ **Orden de impresión:** Apellido paterno A → Z.

⚠️ Si el archivo tiene más de 300 registros, será bloqueado porque probablemente corresponde a un histórico y no al mes actual.
"""
)

uploaded = st.file_uploader("Sube el Excel mensual (.xlsx)", type=["xlsx"])

if uploaded:
    try:
        file_bytes = uploaded.getvalue()
        registros_preview, rbd, periodo = leer_registros(file_bytes)
        total_preview = len(registros_preview)

        st.success("Archivo cargado correctamente.")

        col1, col2, col3 = st.columns(3)
        col1.metric("RBD detectado", rbd if rbd else "No detectado")
        col2.metric("Período", periodo if periodo else "No detectado")
        col3.metric("Docentes encontrados", total_preview)

        if total_preview > MAX_REGISTROS:
            st.error(
                f"Se detectaron {total_preview} registros. "
                "Este archivo parece ser histórico y no mensual. "
                "Por seguridad, no se generará el Excel."
            )
        else:
            st.info("Revise los datos detectados. Si son correctos, genere el Excel.")

            if st.button("Generar Excel imprimible"):
                with st.spinner("Generando archivo..."):
                    try:
                        result_bytes, total = procesar(file_bytes)

                        st.success(
                            f"Archivo generado correctamente. Docentes procesados: {total}"
                        )

                        st.download_button(
                            label="⬇️ Descargar BRP_imprimible.xlsx",
                            data=result_bytes,
                            file_name="BRP_imprimible.xlsx",
                            mime=(
                                "application/vnd.openxmlformats-officedocument."
                                "spreadsheetml.sheet"
                            ),
                        )

                    except Exception as e:
                        st.error(f"Error: {e}")

    except Exception as e:
        st.error(f"Error: {e}")
